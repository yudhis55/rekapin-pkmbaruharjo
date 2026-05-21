"""Excel workbook builder for rekap harian.

Generates Excel matching sample_rekap.xlsx template:
- 2-row header with merged cells
- Dynamic columns: Ruangan (per ruang visited) | Laboratorium (lab treatments) | Tindakan (biasa)
- 1 row per patient (multiple visits merged)
- Jumlah column with SUM formula
- JUMLAH row at bottom with SUM formulas per column
- Kwitansi column empty (manual fill)
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import PatientVisit

# Ruang groups: these ruang names get merged into one column
RUANG_GROUP_KIA_KB_MTBS = {"POLI KIA", "POLI KB", "POLI MTBS"}
RUANG_GROUP_LABEL = "KIA/KB/MTBS"
RUANG_VISIT_FEE = 10000  # Fixed fee per ruang visit
TINDAKAN_RUANGAN_EXCLUDE = {"pelayanan rawat jalan"}


def _normalize_ruang_name(ruang: str) -> str:
    """Normalize ruang name for column header.

    - Strip 'POLI ' prefix
    - Group KIA/KB/MTBS
    - Title case
    """
    ruang_upper = ruang.strip().upper()
    if ruang_upper in RUANG_GROUP_KIA_KB_MTBS:
        return RUANG_GROUP_LABEL
    # Strip POLI prefix
    if ruang_upper.startswith("POLI "):
        name = ruang_upper[5:].strip()
    else:
        name = ruang_upper
    return name.title()


def _get_column_groups(
    visits: list[PatientVisit],
) -> tuple[list[str], list[str], list[str]]:
    """Determine dynamic column groups from visit data.

    Returns:
        (ruang_cols, lab_cols, tindakan_cols) - ordered lists of column names
    """
    ruang_set: set[str] = set()
    lab_set: set[str] = set()
    tindakan_set: set[str] = set()

    for visit in visits:
        ruang_set.add(_normalize_ruang_name(visit.ruang))
        for t in visit.treatments:
            if t.kategori == "lab":
                lab_set.add(t.nama_tindakan)
            else:
                if t.nama_tindakan.lower().strip() not in TINDAKAN_RUANGAN_EXCLUDE:
                    tindakan_set.add(t.nama_tindakan)

    # Sort for consistent ordering; Umum first, KIA/KB/MTBS next, then alphabetical
    ruang_cols = sorted(ruang_set, key=lambda x: (x != "Umum", x != RUANG_GROUP_LABEL, x))
    lab_cols = sorted(lab_set)
    tindakan_cols = sorted(tindakan_set)

    return ruang_cols, lab_cols, tindakan_cols


def _group_visits_by_patient(
    visits: list[PatientVisit],
) -> list[dict[str, Any]]:
    """Group multiple visits per patient into one row.

    Returns list of patient dicts with:
    - no_rm, nama, tgl_lahir, tanggal_kunjungan
    - ruang: normalized ruang name from FIRST visit
    - lab_treatments: dict {nama_tindakan: biaya}
    - biasa_treatments: dict {nama_tindakan: biaya}
    """
    patients: dict[str, dict[str, Any]] = {}

    for visit in visits:
        no_rm = visit.no_rm
        if no_rm not in patients:
            patients[no_rm] = {
                "no_rm": no_rm,
                "nama": visit.nama,
                "tgl_lahir": visit.tgl_lahir,
                "tanggal_kunjungan": visit.tanggal_kunjungan,
                "ruang": _normalize_ruang_name(visit.ruang),
                "lab_treatments": {},
                "biasa_treatments": {},
            }

        p = patients[no_rm]

        for t in visit.treatments:
            if t.kategori == "lab":
                # If same tindakan appears multiple times, sum the biaya
                p["lab_treatments"][t.nama_tindakan] = (
                    p["lab_treatments"].get(t.nama_tindakan, Decimal("0")) + t.biaya
                )
            else:
                if t.nama_tindakan.lower().strip() not in TINDAKAN_RUANGAN_EXCLUDE:
                    p["biasa_treatments"][t.nama_tindakan] = (
                        p["biasa_treatments"].get(t.nama_tindakan, Decimal("0")) + t.biaya
                    )

    return list(patients.values())


def build_rekap_workbook(
    visits: list[PatientVisit],
    tanggal: date,
    cara_bayar: str = "UMUM",
) -> Workbook:
    """Build Excel workbook matching sample_rekap.xlsx template.

    Args:
        visits: List of PatientVisit objects with treatments loaded
        tanggal: Date of the rekap
        cara_bayar: Filter label for filename context

    Returns:
        openpyxl Workbook ready to save
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rekap {tanggal.strftime('%d-%m-%Y')}"

    # Get dynamic column groups
    ruang_cols, lab_cols, tindakan_cols = _get_column_groups(visits)

    # Column layout:
    # A=No, B=Tanggal, C=Nama, D=No RM
    # E...(E+len(ruang_cols)-1) = Ruangan
    # next...(next+len(lab_cols)-1) = Laboratorium
    # next...(next+len(tindakan_cols)-1) = Tindakan
    # Jumlah, Kwitansi

    fixed_cols = 4  # No, Tanggal, Nama, No RM
    ruang_start = fixed_cols + 1  # column E = 5
    ruang_end = ruang_start + len(ruang_cols) - 1 if ruang_cols else fixed_cols

    lab_start = ruang_end + 1
    lab_end = lab_start + len(lab_cols) - 1 if lab_cols else ruang_end

    tindakan_start = lab_end + 1
    tindakan_end = (
        tindakan_start + len(tindakan_cols) - 1 if tindakan_cols else lab_end
    )

    jumlah_col = tindakan_end + 1
    kwitansi_col = jumlah_col + 1

    # ---- ROW 1: Main headers ----
    # Fixed headers (merged rows 1-2)
    for col, header in [(1, "No"), (2, "Tanggal"), (3, "Nama"), (4, "No RM")]:
        ws.cell(row=1, column=col, value=header)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    # Ruangan group header
    if ruang_cols:
        ws.cell(row=1, column=ruang_start, value="Ruangan")
        if len(ruang_cols) > 1:
            ws.merge_cells(
                start_row=1, start_column=ruang_start, end_row=1, end_column=ruang_end,
            )

    # Laboratorium group header
    if lab_cols:
        ws.cell(row=1, column=lab_start, value="Laboratorium")
        if len(lab_cols) > 1:
            ws.merge_cells(
                start_row=1, start_column=lab_start, end_row=1, end_column=lab_end,
            )

    # Tindakan group header
    if tindakan_cols:
        ws.cell(row=1, column=tindakan_start, value="Tindakan")
        if len(tindakan_cols) > 1:
            ws.merge_cells(
                start_row=1,
                start_column=tindakan_start,
                end_row=1,
                end_column=tindakan_end,
            )

    # Jumlah and Kwitansi (merged rows 1-2)
    ws.cell(row=1, column=jumlah_col, value="Jumlah")
    ws.merge_cells(
        start_row=1, start_column=jumlah_col, end_row=2, end_column=jumlah_col,
    )
    ws.cell(row=1, column=kwitansi_col, value="Kwitansi")
    ws.merge_cells(
        start_row=1, start_column=kwitansi_col, end_row=2, end_column=kwitansi_col,
    )

    # ---- ROW 2: Sub-headers ----
    for i, ruang in enumerate(ruang_cols):
        ws.cell(row=2, column=ruang_start + i, value=ruang)
    for i, lab in enumerate(lab_cols):
        ws.cell(row=2, column=lab_start + i, value=lab)
    for i, tindakan in enumerate(tindakan_cols):
        ws.cell(row=2, column=tindakan_start + i, value=tindakan)

    # ---- DATA ROWS ----
    patients = _group_visits_by_patient(visits)
    data_start_row = 3

    for idx, patient in enumerate(patients, start=1):
        row = data_start_row + idx - 1

        # Fixed columns
        ws.cell(row=row, column=1, value=idx)
        tanggal_str = patient["tanggal_kunjungan"].strftime("%d/%m/%Y")
        ws.cell(row=row, column=2, value=tanggal_str)
        ws.cell(row=row, column=3, value=patient["nama"])
        ws.cell(row=row, column=4, value=patient["no_rm"])

        # Ruangan columns: 10000 per ruang visited
        for i, ruang in enumerate(ruang_cols):
            if ruang == patient["ruang"]:
                ws.cell(row=row, column=ruang_start + i, value=RUANG_VISIT_FEE)

        # Laboratorium columns
        for i, lab in enumerate(lab_cols):
            biaya = patient["lab_treatments"].get(lab)
            if biaya and biaya > 0:
                ws.cell(row=row, column=lab_start + i, value=int(biaya))

        # Tindakan columns
        for i, tindakan in enumerate(tindakan_cols):
            biaya = patient["biasa_treatments"].get(tindakan)
            if biaya and biaya > 0:
                ws.cell(row=row, column=tindakan_start + i, value=int(biaya))

        # Jumlah formula: sum all biaya columns in this row
        first_biaya_col = get_column_letter(ruang_start)
        last_biaya_col = get_column_letter(jumlah_col - 1)
        ws.cell(
            row=row,
            column=jumlah_col,
            value=f"=SUM({first_biaya_col}{row}:{last_biaya_col}{row})",
        )

        # Kwitansi: empty (manual fill)

    # ---- JUMLAH ROW ----
    jumlah_row = data_start_row + len(patients)
    ws.cell(row=jumlah_row, column=2, value="JUMLAH")

    # Sum formulas for each biaya column (including the Jumlah column)
    if patients:
        for col in range(ruang_start, jumlah_col + 1):
            col_letter = get_column_letter(col)
            ws.cell(
                row=jumlah_row,
                column=col,
                value=f"=SUM({col_letter}{data_start_row}:{col_letter}{jumlah_row - 1})",
            )

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize workbook to bytes for HTTP response."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
