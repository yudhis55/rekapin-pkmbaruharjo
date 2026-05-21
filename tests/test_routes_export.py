"""Tests for Excel export route."""
from __future__ import annotations

from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture(autouse=True)
def _setup(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> Iterator[None]:
    monkeypatch.setenv("EMR_USERNAME", "u")
    monkeypatch.setenv("EMR_PASSWORD", "p")
    monkeypatch.setenv("EMR_BASE_URL", "https://example.test/daf")
    monkeypatch.setenv("EMR_PUSKESMAS", "P")
    db_file = tmp_path / "t.db"
    monkeypatch.setenv("DATABASE_URL", f"sqlite+aiosqlite:///{db_file}")
    from config.settings import get_settings
    get_settings.cache_clear()
    import app.db.session as db_mod
    db_mod._engine = None
    db_mod._sessionmaker = None
    yield


async def _seed(client: TestClient) -> None:
    """Seed database with sample visits via direct repo calls."""
    from app.db.repositories import visit_repo
    from app.db.session import get_sessionmaker
    sm = get_sessionmaker()
    async with sm() as session:
        await visit_repo.upsert_visit(
            session,
            visit_repo.VisitInput(
                emr_visit_id="EMR001",
                no_rm="RM001", nama="Sample 1", tgl_lahir=date(1990, 1, 1),
                ruang="POLI UMUM", tanggal_kunjungan=date(2026, 5, 15),
                cara_bayar="UMUM",
                total_biaya=Decimal("0.00"),
            ),
            [],
        )
        await visit_repo.upsert_visit(
            session,
            visit_repo.VisitInput(
                emr_visit_id="EMR002",
                no_rm="RM002", nama="Sample 2", tgl_lahir=None,
                ruang="POLI GIGI", tanggal_kunjungan=date(2026, 5, 15),
                cara_bayar="BPJS",
                total_biaya=Decimal("0.00"),
            ),
            [],
        )


def test_export_excel_returns_xlsx() -> None:
    import asyncio

    from app.main import create_app
    app = create_app()
    with TestClient(app) as client:
        asyncio.get_event_loop().run_until_complete(_seed(client))
        r = client.get("/api/export/excel", params={"tanggal": "2026-05-15"})
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]


def test_export_excel_filename_format() -> None:
    import asyncio

    from app.main import create_app
    app = create_app()
    with TestClient(app) as client:
        asyncio.get_event_loop().run_until_complete(_seed(client))
        r = client.get("/api/export/excel", params={"tanggal": "2026-05-15"})
        assert r.status_code == 200
        cd = r.headers["content-disposition"]
        assert "rekap_2026-05-15_SEMUA.xlsx" in cd


def test_export_excel_empty_returns_empty_workbook() -> None:
    from app.main import create_app
    app = create_app()
    with TestClient(app) as client:
        r = client.get("/api/export/excel", params={"tanggal": "2099-01-01"})
        assert r.status_code == 200
        body = r.content
        assert len(body) > 0
        # Verify it's a valid xlsx file
        wb = load_workbook(BytesIO(body))
        assert wb.active is not None


def test_export_excel_filter_cara_bayar() -> None:
    import asyncio

    from app.main import create_app
    app = create_app()
    with TestClient(app) as client:
        asyncio.get_event_loop().run_until_complete(_seed(client))
        r = client.get(
            "/api/export/excel",
            params={"tanggal": "2026-05-15", "cara_bayar": "UMUM"},
        )
        assert r.status_code == 200
        body = r.content
        assert len(body) > 0
        wb = load_workbook(BytesIO(body))
        assert wb.active is not None
        cd = r.headers["content-disposition"]
        assert "rekap_2026-05-15_UMUM.xlsx" in cd
